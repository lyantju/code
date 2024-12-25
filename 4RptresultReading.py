import os
from openpyxl import Workbook

# Get the path of the folder containing the .rpt files
folder_path = 'D:\WRR\swmm'# Replace with your folder path
file_names = [f for f in os.listdir(folder_path) if f.endswith('.rpt')]

wb = Workbook()

# Loop through each .rpt file
for file_name in file_names:
    file_path = os.path.join(folder_path, file_name)
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
        lines = file.readlines()

    # Extract the specified lines (line numbers are inclusive of start and end)
    lines_to_write = lines[24544:26555] # Update these indices if needed

    sheet_name = os.path.splitext(file_name)[0]
    ws = wb.create_sheet(title=sheet_name)

    for i, line in enumerate(lines_to_write, start=1):
        ws.cell(row=i, column=1, value=line.strip())

wb.remove(wb.active)

excel_file_path = os.path.join(folder_path, 'Drainagepipeinformation.xlsx')
wb.save(excel_file_path)